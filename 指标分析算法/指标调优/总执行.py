from 指标对比_除D0 import zhibiao
import openpyxl
def zhixing(diqv,time):
    diqv = diqv
    time =time
    Alllist = []
    print("开始")
    if diqv == "D0":
        size_list = [0.2,0.3,0.4]
    elif diqv == "D1":
        size_list = [0.2,0.3,0.4,0.5]
    else:
        size_list = [0.3,0.4,0.5]
    for i in size_list:
        for j in range(1,60):
            test_size,random_state,R2_improvement, _ = zhibiao(i,j,diqv,time)
            if R2_improvement == 0:
                continue
            element = {
                "test_size":test_size,
                "random_state":random_state,
                "R2_improvement":R2_improvement,
            }
            Alllist.append(element)
    if len(Alllist) == 0:
        print("没有满足条件的字典")
        print("结束")
        return 0,0,0
    else:
        print("Alllist长度:", len(Alllist))
        max_item = max(Alllist, key=lambda d: d["R2_improvement"])
        print("最大 R2_improvement的字典:", max_item)
        max_R2_improvement = max_item["R2_improvement"]
        print("最大 R2_improvement:", max_R2_improvement)
        test_size,random_state, _, all_data = zhibiao(max_item["test_size"],max_item["random_state"],diqv,time)
        print("结束")
        return test_size,random_state,all_data

if __name__ == '__main__':
    file_lujing = r"F:/xiangmu/交通指标计算/程师兄数据处理/模型性能比较.xlsx"
    # 加载已有工作簿
    wb = openpyxl.load_workbook(file_lujing)
    for sheet in wb.sheetnames:
        print(sheet)
    ws = wb['模型性能比较']
    # 定义起始行和起始列（Excel中行和列均从1开始计数）
    start_row = 0
    start_col = 0
    hang = 1
    for p in ["D0"]: #"D0","D1","D2","D3","D4","D5","D6"
        start_col = start_col + 3
        for q in ["Labour Day","Have to work","workday","weekend"]:#,"Have to work","workday","weekend"
            print(p + q + "开始执行")
            if start_row >= 12:
                start_row = 0
            start_row = start_row + 3
            hang = hang + 1
            # 要写入的数据列表
            test_size,random_state,data = zhixing(p,q)
            if data == 0 and random_state == 0 and test_size == 0:
                continue
            # 使用嵌套的列表推导式来保留四位小数
            data = [[round(num, 4) for num in row] for row in data]
            data[2] = [f"{x}%" for x in data[2]]
            print(data)
            # 遍历数据，并更新指定单元格
            for m, row in enumerate(data):
                for n, value in enumerate(row):
                    print(m)
                    print(n)
                    print(value)
                    ws.cell(row=start_row + m, column=start_col + n, value=value)
                    # 保存更改
                    wb.save(file_lujing)
            import subprocess
            # from SHAP import compute_metrics
            # compute_metrics(i,j,hang,test_size,random_state)
            result = subprocess.run(
                [r"F:/Anconda/envs/shap_env/python.exe", "SHAP.py", str(p),str(q),str(hang), str(test_size), str(random_state)],
                capture_output=True,
                text=True,
                encoding="utf-8"
            )
            # 输出标准输出
            print("标准输出:")
            print(result.stdout)
            print(f"标准错误输出: {result.stderr}")

            print(p + q + "已完成")


